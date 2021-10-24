import random
all=["許諺謹","孫健崴","陳隆瑞","沈泰暘","馬仁佐","柯柏安","游鎛謙","林鈺笙","高信智","李梓霆","李文勝","郭子儀","楊子興","莫明翰","張庭瑞","張容碩","王大謙","林晉德","許博能"]
noForT=["許諺謹","孫健崴","柯柏安","游鎛謙","李文勝","張容碩","林晉德"]
Flag_Teach=['楊子興', '沈泰暘', '陳隆瑞', '馬仁佐', '林鈺笙', '郭子儀', '許博能', '莫明翰', '張庭瑞', '王大謙', '高信智', '李梓霆']
# print(len(all)) 
# 19
def get_ran_honorguard ():
    array=all
    random.shuffle(array)
    return array

from openpyxl import load_workbook

excel=load_workbook("ran.xlsx")
sheet=excel.active
everyone=[]

for row in range(1,31):
    for col in range(1,4):
        take=((row-1)*3+(col-1))%19
        if not take:
            everyone=get_ran_honorguard()
        sheet.cell(row,col,everyone[take])



excel.save("ran.xlsx")

# To check

for row in range(1,31):
    for col in range(1,4):
        print(sheet.cell(row,col).value,end=" ")
    print("")